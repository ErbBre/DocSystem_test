import openpyxl

# Ruta al archivo Excel
file_path = 'APP/MODEL/excel_files/Libro_dcs.xlsx'

# Cargar el libro de trabajo (workbook) desde la ruta especificada
workbook = openpyxl.load_workbook(file_path)

# Función para imprimir los nombres de las hojas en el libro
def print_sheets(workbook):
    sheets = workbook.sheetnames
    print("Hojas en el libro:")
    for sheet in sheets:
        print(sheet)

# Función para leer el valor de una celda específica
def read_cell(sheet, cell):
    return sheet[cell].value

# Función para concatenar las palabras en orden de columnas y formar una oración
def concatenate_sentence(sheet):
    words = []
    # Iterar sobre las columnas (de izquierda a derecha)
    for col in sheet.iter_cols(min_row=1, max_col=8, max_row=12, values_only=True):
        # Iterar sobre las celdas en cada columna
        for cell in col:
            if cell:  # Si la celda no está vacía
                words.append(str(cell))  # Agregar el valor de la celda a la lista de palabras
    # Unir las palabras con espacios para formar la oración
    sentence = ' '.join(words)
    return sentence

# Seleccionar la hoja activa del libro
sheet = workbook.active

# Llamar a la función para concatenar las palabras en orden de columnas
sentence = concatenate_sentence(sheet)

# Imprimir la oración concatenada
print(f"Oración concatenada: {sentence}")