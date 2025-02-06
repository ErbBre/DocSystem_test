import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border,Side

# Create a new workbook
new_workbook = openpyxl.Workbook()

# Get the active sheet (default sheet when you create a new workbook)
new_sheet = new_workbook.active
new_sheet.title = "Parents"
# Definir el estilo de los bordes
borde = Border(
    top     = Side(border_style='thin', color='FFFFFF'),  # Borde superior
    bottom  = Side(border_style='thin', color='FFFFFF'),  # Borde inferior
    left    = Side(border_style='thin', color='FFFFFF'),  # Borde izquierdo
    right   = Side(border_style='thin', color='FFFFFF')   # Borde derecho
)
# HEAD OF REPORT
title_report = {"A":"CODIGO",
                "B":"NOMBRES Y APELLIDOS",
                "C":"NUM. DOCUMENTO",
                "D":"FECHA DE NACIMIENTO",
                "E":"GENERO",
                "F":"OCUPACIÓN",
                "G":"TELEFONO",
                "H":"CORREO",
                "I":"ESTADO"}
new_sheet.merge_cells('A1:I1')

new_sheet['A1'] = "REPORT OF PARENTS"
new_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
new_sheet['A1'].font = Font(bold=True, color="FFFFFF")
new_sheet['A1'].fill = PatternFill(start_color="4e0081", end_color="4e0081", fill_type="solid")

for col1, col2 in title_report.items():
    new_sheet[f'{col1}2'] = col2

    # Change font style
    new_sheet[f'{col1}2'].font = Font(bold=True, color="FFFFFF")
    # Set cell alignment
    new_sheet[f'{col1}2'].alignment = Alignment(horizontal='center', vertical='center')
    # Set background color
    new_sheet[f'{col1}2'].fill = PatternFill(start_color="4e0081", end_color="4e0081", fill_type="solid")

    new_sheet[f'{col1}2'].border = borde
    new_sheet.row_dimensions[1].height = 25 
    new_sheet.row_dimensions[2].height = 25 




# Create a new sheet
new_sheet = new_workbook.create_sheet("Sheet2")

# Access a specific sheet
sheet2 = new_workbook["Sheet2"]
# Save the new workbook
new_workbook.save('./APP/MODEL/excel_files/new_file.xlsx')
